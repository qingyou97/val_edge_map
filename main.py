import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage

# 创建Excel工作簿和工作表
wb = Workbook()
ws = wb.active

# 设置文件夹路径
folder_path = 'A'

# 设置图像大小
img_width = 30
img_height = 30

# 遍历文件夹
for row, folder_name in enumerate(os.listdir(folder_path), start=1):
    folder_path_full = os.path.join(folder_path, folder_name)
    if os.path.isdir(folder_path_full):
        # 在指定单元格写上文件夹名
        ws.cell(row=row, column=1, value=folder_name)

        # 获取文件夹中的图像文件
        img_files = [f for f in os.listdir(folder_path_full) if f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp'))]
        
        for col, img_filename in enumerate(img_files, start=2):
            img_path = os.path.join(folder_path_full, img_filename)

            # 使用Pillow来改变图像大小
            img = PILImage.open(img_path)
            img = img.resize((img_width, img_height), PILImage.LANCZOS)
            resized_img_path = os.path.join(folder_path_full, f'resized_{img_filename}')
            img.save(resized_img_path)

            # 打开重新调整大小后的图像
            openpyxl_img = Image(resized_img_path)
            # 将图像插入到Excel中
            ws.add_image(openpyxl_img, ws.cell(row=row, column=col).coordinate)

            # 设置列宽为图像宽度
            col_letter = ws.cell(row=row, column=col).column_letter
            ws.column_dimensions[col_letter].width = img_width / 7.5  # Excel列宽单位与像素不同，需转换

        # 设置行高为图像高度
        ws.row_dimensions[row].height = img_height

# 保存Excel文件
wb.save('output.xlsx')
